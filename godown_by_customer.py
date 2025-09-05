import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from datetime import datetime

# ---------------- FUNCTION ----------------
def split_excel_by_customer(uploaded_file, selected_customer):
    # Read Excel
    df = pd.read_excel(uploaded_file, sheet_name=0, dtype=str)

    # Clean column headers
    df.columns = df.columns.str.strip()

    # Rename Item Code -> ZFI Part No
    if "Item Code" in df.columns:
        df = df.rename(columns={"Item Code": "ZFI Part No"})

    # Remove rows starting with C/c in ZFI Part No
    if "ZFI Part No" in df.columns:
        df = df[~df["ZFI Part No"].str.startswith(("C", "c"), na=False)]

    # Convert Qty & Amount to numeric
    if "Qty" in df.columns:
        df["Qty"] = pd.to_numeric(df["Qty"], errors="coerce").fillna(0).astype(float)
    if "Amount" in df.columns:
        df["Amount"] = pd.to_numeric(df["Amount"], errors="coerce").fillna(0).astype(float)

    # Smart grouping key: Customer Part No if present, else ZFI Part No
    if "Customer Part No" in df.columns and "ZFI Part No" in df.columns:
        df["Group Key"] = df["Customer Part No"].fillna("").str.strip()
        df["Group Key"] = df["Group Key"].mask(df["Group Key"] == "", df["ZFI Part No"])
    else:
        df["Group Key"] = df["ZFI Part No"]

    # Clean customer names
    if "Name" in df.columns:
        df["Name"] = (
            df["Name"]
            .fillna("Unknown Customer")
            .str.replace(r"\s+", " ", regex=True)
            .str.strip()
        )

    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    total_fill = PatternFill("solid", fgColor="D9D9D9")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    required_cols = ["Customer Part No", "ZFI Part No", "Item Desc",
                     "Inv No", "Inv Date", "Qty", "Amount"]

    # If "All Customers" selected -> process all
    if selected_customer == "All Customers":
        customer_groups = df.groupby("Name")
    else:
        customer_groups = [(selected_customer, df[df["Name"] == selected_customer])]

    # Create sheet(s)
    for cust_name, group in customer_groups:
        ws = wb.create_sheet(title=str(cust_name)[:31])  # sheet name max 31 chars
        ws.append(required_cols)

        # Header styling
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

        row_idx = 2

        # Group by smart key
        for key, part_group in group.groupby("Group Key"):
            sum_qty = part_group["Qty"].sum()
            sum_amount = part_group["Amount"].sum()

            first_row = True
            for _, r in part_group.iterrows():
                ws.append([
                    r.get("Customer Part No", "") if first_row else "",
                    r.get("ZFI Part No", "") if first_row else "",
                    r.get("Item Desc", ""),
                    r.get("Inv No", ""),
                    r.get("Inv Date", ""),
                    r.get("Qty", 0),
                    r.get("Amount", 0)
                ])
                first_row = False
                row_idx += 1

            # Add subtotal row for the group
            ws.append([f"{key} Total", "", "", "", "", sum_qty, sum_amount])
            for cell in ws[row_idx]:
                cell.font = Font(bold=True)
                cell.fill = total_fill
                cell.border = thin_border
            row_idx += 1

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2

    # Save workbook to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output

# ---------------- STREAMLIT APP ----------------
st.title("📊 Split Excel by Customer Name with Subtotals")

uploaded_file = st.file_uploader("Upload your Excel file", type=["xlsx"])

if uploaded_file:
    # Preview available customers
    df_preview = pd.read_excel(uploaded_file, sheet_name=0, dtype=str)
    df_preview.columns = df_preview.columns.str.strip()

    if "Name" in df_preview.columns:
        df_preview["Name"] = (
            df_preview["Name"]
            .fillna("Unknown Customer")
            .str.replace(r"\s+", " ", regex=True)
            .str.strip()
        )
        customer_list = sorted(df_preview["Name"].unique().tolist())
        customer_list.insert(0, "All Customers")  # add option at top

        selected_customer = st.selectbox("Select Customer", customer_list)

        if st.button("Process File"):
            processed_file = split_excel_by_customer(uploaded_file, selected_customer)

            # Build dynamic file name
            today_str = datetime.today().strftime("%d.%m.%Y")
            download_filename = f"{selected_customer} godown stock dt {today_str}.xlsx"

            st.download_button(
                label="📥 Download Formatted Excel",
                data=processed_file,
                file_name=download_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    else:
        st.error("❌ 'Name' column (Customer Name) not found in Excel file!")
